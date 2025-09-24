from __future__ import annotations

"""excel.inspect tool skeleton (static preflight of Excel workbooks).

Security posture:
- No macro execution. Read-only parsing.
- Intended to operate on RO mounts, emit artifacts to RW/audit mounts.

MVP: This is a non-functional scaffold. Implementation will:
- Parse workbook inventory (sheets, named ranges, formulas)
- Detect volatile functions
- Enumerate external links/connections
- Extract VBA modules (no execution)
- Write preflight.json/md and artifact hashes
"""

from typing import Dict, Any, Generator, Tuple, List
from pathlib import Path
import hashlib
import json
import os
import time
import zipfile

# Optional deps
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - optional
    load_workbook = None  # type: ignore
try:
    from oletools.olevba import VBA_Parser  # type: ignore
except Exception:  # pragma: no cover - optional
    VBA_Parser = None  # type: ignore
try:
    import msoffcrypto  # type: ignore
except Exception:  # pragma: no cover - optional
    msoffcrypto = None  # type: ignore

from ...mcp.runtime import JOB_MANAGER
from ...deps import get_status


VOLATILE_FUNCS = (
    "INDIRECT", "OFFSET", "NOW", "TODAY", "RAND", "RANDBETWEEN", "INFO", "CELL", "AREAS",
)


def _sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open('rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def _safe_rel(base: Path, target: Path) -> str:
    try:
        return str(target.resolve().relative_to(base.resolve()))
    except Exception:
        return str(target)


def stream(job_id: str, params: Dict[str, Any]) -> Generator[Dict[str, Any], None, None]:
    """Generate SSE-friendly events for excel.inspect.

    MVP implementation:
    - Validates path
    - For .xlsx/.xlsm: zip scan for parts, vbaProject.bin, worksheets, connections, externalLinks
    - Writes minimal preflight.json/md and computes hashes
    - Emits progress events and a final result
    """
    t0 = time.time()
    budget_s = int(params.get('budget_s', 120))
    warned_80 = False
    path_s = str(params.get('path') or '')
    lint = bool(params.get('lint', True))
    extract_vba = bool(params.get('extract_vba', True))
    password = params.get('password')

    if not path_s:
        yield {"event": "fault", "job": job_id, "ok": False, "error": "path missing"}
        return
    src = Path(path_s)
    if not src.exists() or not src.is_file():
        yield {"event": "fault", "job": job_id, "ok": False, "error": "input not found or not a file"}
        return
    # Enforce absolute path and RO root (best-effort)
    try:
        real = src.resolve()
        ro = Path('/mnt/ro').resolve()
        if not str(real).startswith(str(ro)):
            yield {"event": "fault", "job": job_id, "ok": False, "error": "path must be under /mnt/ro"}
            return
    except Exception:
        pass

    # Prepare output dirs
    out_root = Path('out') / 'tools' / job_id
    vba_dir = out_root / 'vba'
    parts_dir = out_root / 'parts'
    out_root.mkdir(parents=True, exist_ok=True)

    yield {"event": "progress", "job": job_id, "pct": 10, "msg": "Opening workbook"}
    if (time.time() - t0) > 0.8 * budget_s and not warned_80:
        yield {"event": "warning", "job": job_id, "msg": "Approaching time budget (80%)"}
        warned_80 = True

    # Basic type detection
    suffix = src.suffix.lower()
    kind = 'unknown'
    has_vba = False
    ws_count = 0
    external_links: List[str] = []
    connections_present = False
    notes: List[str] = []
    total_formulas = 0
    volatile_hits: Dict[str, int] = {k: 0 for k in VOLATILE_FUNCS}
    arrays = 0
    refs_3d = 0

    if suffix in ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xlsb'):
        kind = 'openxml'
        try:
            with zipfile.ZipFile(src, 'r') as z:
                # Zip-bomb guard (soft)
                names = z.namelist()
                if len(names) > 10000:
                    yield {"event": "fault", "job": job_id, "ok": False, "error": "zip too large (members > 10000)"}
                    return
                total_uncomp = 0
                for info in z.infolist():
                    total_uncomp += getattr(info, 'file_size', 0)
                    if total_uncomp > 500 * 1024 * 1024:  # 500MB
                        yield {"event": "fault", "job": job_id, "ok": False, "error": "zip too large (>500MB)"}
                        return
                # Encrypted OOXML package (basic detection)
                if 'EncryptionInfo' in names:
                    if not password:
                        yield {"event": "fault", "job": job_id, "ok": False, "error": "PASSWORD_REQUIRED"}
                        return
                    if msoffcrypto is None:
                        yield {"event": "fault", "job": job_id, "ok": False, "error": "PASSWORD_PROVIDED_BUT_MSOFFCRYPTO_MISSING"}
                        return
                    dec_path = (out_root / 'decrypted.xlsx').resolve()
                    try:
                        with open(src, 'rb') as fsrc:
                            of = msoffcrypto.OfficeFile(fsrc)  # type: ignore
                            of.load_key(password=password)
                            with open(dec_path, 'wb') as fdst:
                                of.decrypt(fdst)
                        src = Path(dec_path)
                    except Exception:
                        yield {"event": "fault", "job": job_id, "ok": False, "error": "PASSWORD_INVALID"}
                        return

            # Re-open (possibly decrypted) and enumerate parts
            with zipfile.ZipFile(src, 'r') as z2:
                names = z2.namelist()
                yield {"event": "progress", "job": job_id, "pct": 30, "msg": "Enumerating sheets & parts"}
                ws_count = sum(1 for n in names if n.startswith('xl/worksheets/sheet') and n.endswith('.xml'))
                connections_present = 'xl/connections.xml' in names
                external_links = [n for n in names if n.startswith('xl/externalLinks/') and n.endswith('.xml')]
                has_vba = any(n.endswith('vbaProject.bin') for n in names)

                # Extract selected parts for audit
                parts_dir.mkdir(parents=True, exist_ok=True)
                for p in ['[Content_Types].xml', 'xl/workbook.xml', 'xl/_rels/workbook.xml.rels', 'xl/connections.xml']:
                    if p in names:
                        try:
                            z2.extract(p, parts_dir)
                        except Exception:
                            pass
                for n in external_links[:5]:  # limit for MVP
                    try:
                        z2.extract(n, parts_dir)
                    except Exception:
                        pass

                yield {"event": "progress", "job": job_id, "pct": 65, "msg": "Extracting VBA modules"}
                vba_bins: List[Path] = []
                if extract_vba and has_vba:
                    vba_dir.mkdir(parents=True, exist_ok=True)
                    for n in names:
                        if n.endswith('vbaProject.bin'):
                            try:
                                outp = vba_dir / n
                                outp.parent.mkdir(parents=True, exist_ok=True)
                                z2.extract(n, vba_dir)
                                try:
                                    os.chmod((vba_dir / n).resolve(), 0o640)
                                except Exception:
                                    pass
                                vba_bins.append((vba_dir / n).resolve())
                            except Exception:
                                pass
                if suffix == '.xlsb':
                    # Optional: count sheets via pyxlsb (non-fatal if missing)
                    try:
                        from pyxlsb import open_workbook as xlsb_open  # type: ignore
                        with xlsb_open(str(src)) as wb2:  # pragma: no cover - optional
                            try:
                                ws_count = len(list(wb2.sheets))  # type: ignore[attr-defined]
                            except Exception:
                                pass
                    except Exception:
                        pass
                    notes.append('xlsb formulas limited; values only via pyxlsb')
        except zipfile.BadZipFile:
            yield {"event": "fault", "job": job_id, "ok": False, "error": "invalid zip container"}
            return
    elif suffix in ('.xls',):
        kind = 'biff8'
        notes.append('xls formula text limited; consider converting to xlsx')
        # Optional: count sheets via xlrd 1.2.0 if available
        try:
            import xlrd  # type: ignore
            book = xlrd.open_workbook(str(src), on_demand=True)
            ws_count = book.nsheets
            book.release_resources()
        except Exception:
            notes.append('xlrd not available or failed (non-fatal)')
    elif suffix not in ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xlsb', '.xls'):
        notes.append('unrecognized extension; proceeding with minimal metadata')

    # Optional: scan formulas with openpyxl for xlsx/xlsm
    if load_workbook is not None and suffix in ('.xlsx', '.xlsm'):
        if JOB_MANAGER.is_cancelled(job_id):
            yield {"event": "fault", "job": job_id, "ok": False, "error": "CANCELLED"}
            return
        yield {"event": "progress", "job": job_id, "pct": 55, "msg": "Scanning formulas (openpyxl)"}
        try:
            wb = load_workbook(filename=str(src), data_only=False, read_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        v = getattr(cell, 'value', None)
                        if isinstance(v, str) and v.startswith('='):
                            total_formulas += 1
                            up = v.upper()
                            # Simple 3D ref heuristic: contains '!' and ':' before '!'
                            if '!' in up and up.split('!')[0].count(':') == 1:
                                refs_3d += 1
                            for fn in VOLATILE_FUNCS:
                                if fn in up:
                                    volatile_hits[fn] += 1
                    # Budget/warning checks per row
                    if (time.time() - t0) > 0.8 * budget_s and not warned_80:
                        yield {"event": "warning", "job": job_id, "msg": "Approaching time budget (80%)"}
                        warned_80 = True
                    if (time.time() - t0) > budget_s:
                        notes.append('formula scan truncated due to budget')
                        wb.close()
                        raise RuntimeError('budget_exceeded')
                if JOB_MANAGER.is_cancelled(job_id):
                    yield {"event": "fault", "job": job_id, "ok": False, "error": "CANCELLED"}
                    return
            wb.close()
        except Exception:
            notes.append('openpyxl formula scan failed (non-fatal)')

    # Optional: VBA analysis via oletools (no execution)
    vba_summary = None
    try:
        if 'vba_bins' in locals() and VBA_Parser is not None and vba_bins:
            risky_patterns = (
                'CreateObject', 'GetObject', 'Shell', 'Application.Run', 'ExecuteExcel4Macro', 'Solver', 'ActiveX',
            )
            modules = set()
            procedures = []
            risky_calls = set()
            import re
            proc_re = re.compile(r"^\s*(Sub|Function)\s+([A-Za-z0-9_]+)\s*\(", re.IGNORECASE | re.MULTILINE)
            for vb in vba_bins:
                try:
                    vp = VBA_Parser(str(vb))
                    if vp.detect_vba_macros():
                        for (_, _stream, vba_filename, vba_code) in vp.extract_all_macros():
                            if vba_filename:
                                modules.add(vba_filename)
                            if isinstance(vba_code, str):
                                for m in proc_re.finditer(vba_code):
                                    kind, name = m.group(1), m.group(2)
                                    procedures.append({"module": vba_filename or "", "name": name, "kind": kind})
                                up = vba_code
                                for pat in risky_patterns:
                                    if pat in up:
                                        risky_calls.add(pat)
                except Exception:
                    continue
            vba_summary = {
                "modules": sorted(modules),
                "procedures": procedures[:200],  # cap for MVP
                "risky_calls": sorted(risky_calls),
            }
    except Exception:
        notes.append('oletools VBA analysis failed (non-fatal)')

    yield {"event": "progress", "job": job_id, "pct": 80, "msg": "Computing hashes & writing reports"}

    # Compute hashes and write outputs
    input_sha = _sha256_file(src)
    preflight_json = out_root / 'preflight.json'
    preflight_md = out_root / 'preflight.md'

    flags = {
        "has_macros": bool(has_vba),
        "external_links": len(external_links),
        "connections": bool(connections_present),
        "formulas": int(total_formulas),
        "volatile": {k: v for k, v in volatile_hits.items() if v > 0},
        "arrays": int(arrays),
        "refs3d": int(refs_3d),
    }
    # Simple severity heuristic
    vol_ratio = 0.0
    if total_formulas > 0:
        vol_ratio = sum(flags["volatile"].values()) / float(total_formulas)
    score = (3 if flags["has_macros"] else 0) + (2 if flags["external_links"] else 0) + (1 if flags["connections"] else 0) + (1 if vol_ratio > 0.02 else 0)
    severity = 'green'
    if score >= 4:
        severity = 'red'
    elif score >= 2:
        severity = 'amber'

    payload = {
        "path": str(src),
        "kind": kind,
        "worksheet_count": ws_count,
        "has_vba": has_vba,
        "external_links": external_links,
        "connections": connections_present,
        "notes": notes,
        "flags": {
            "severity": severity,
            "issues": score,
            "breakdown": {
                "has_macros": bool(has_vba),
                "external_links": len(external_links) > 0,
                "connections": bool(connections_present),
                "volatile_ratio": vol_ratio,
                "refs3d": refs_3d,
            }
        },
        "input_sha256": input_sha,
        "created_at": int(time.time()),
        "tool": "excel.inspect",
        "params": {"lint": lint, "extract_vba": extract_vba, "password": bool(password), "budget_s": budget_s},
        "mode": str(get_status().mode),
    }
    if vba_summary is not None:
        payload["vba"] = vba_summary
    preflight_json.write_text(json.dumps(payload, indent=2), encoding='utf-8')
    pre_md = [
        f"# Excel Preflight: {src.name}",
        "", f"- Kind: {kind}", f"- Sheets: {ws_count}", f"- Macros: {'yes' if has_vba else 'no'}",
        f"- External links: {len(external_links)}", f"- Connections: {'yes' if connections_present else 'no'}",
        f"- Severity: {severity}",
    ]
    if notes:
        pre_md += ["", "## Notes"] + [f"- {n}" for n in notes]
    preflight_md.write_text("\n".join(pre_md) + "\n", encoding='utf-8')

    # Hash artifacts
    artifacts = [preflight_json, preflight_md]
    if vba_dir.exists():
        for p in vba_dir.rglob('*'):
            if p.is_file():
                artifacts.append(p)
    if parts_dir.exists():
        for p in parts_dir.rglob('*'):
            if p.is_file():
                artifacts.append(p)
    artifact_entries = [{"path": str(_safe_rel(out_root, a)), "sha256": _sha256_file(a)} for a in artifacts]
    # job.json and audit.json
    try:
        params_norm = {"path": str(src), "lint": lint, "extract_vba": extract_vba, "password": bool(password), "budget_s": budget_s}
        params_hash = hashlib.sha256(json.dumps(params_norm, sort_keys=True).encode('utf-8')).hexdigest()
        job = {
            "tool": "excel.inspect",
            "job_id": job_id,
            "params_hash": params_hash,
            "started_at": int(t0),
            "completed_at": int(time.time()),
            "ok": True,
            "mode": str(get_status().mode),
            "artifacts": artifact_entries,
        }
        (out_root / 'job.json').write_text(json.dumps(job, indent=2), encoding='utf-8')
        audit_path = Path('out') / 'audit.json'
        try:
            existing = json.loads(audit_path.read_text(encoding='utf-8'))
            if not isinstance(existing, list):
                existing = []
        except Exception:
            existing = []
        existing.append({
            "job_id": job_id,
            "tool": "excel.inspect",
            "input_sha256": input_sha,
            "flags": payload["flags"],
            "artifacts": artifact_entries,
            "ts": int(time.time()),
        })
        audit_path.write_text(json.dumps(existing, indent=2), encoding='utf-8')
    except Exception:
        pass

    yield {
        "event": "result",
        "job": job_id,
        "ok": True,
        "artifact": str(preflight_json),
        "report": str(preflight_md),
        "sha256": input_sha,
        "flags": {"severity": severity, "issues": score},
        "artifacts": artifact_entries,
    }
